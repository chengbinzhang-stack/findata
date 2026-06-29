import os
import uuid
import asyncio
from pathlib import Path
from datetime import date, datetime
from contextlib import asynccontextmanager

from fastapi import FastAPI, Request, Form, HTTPException, BackgroundTasks
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse
from fastapi.templating import Jinja2Templates
from typing import Optional
import httpx

from models import Quarter, DocumentType, TaskStatus, ValidateUrlResponse
from database import (
    init_db, search_companies, get_company_by_scrip, get_company_by_name,
    save_company, create_download_task, get_task, update_task_status,
    create_file_record, update_file_record, get_file_records, get_all_files
)
from downloader import validate_url, generate_filename, generate_folder_path, download_file
from s3_utils import upload_to_s3


DOWNLOAD_DIR = Path(__file__).parent / "downloads"
S3_BUCKET = os.getenv("S3_BUCKET", "")


@asynccontextmanager
async def lifespan(app: FastAPI):
    await init_db()
    DOWNLOAD_DIR.mkdir(exist_ok=True)
    yield


app = FastAPI(title="Transcript Downloader", lifespan=lifespan)
templates = Jinja2Templates(directory="templates")


def get_current_fy_choices():
    current_year = datetime.now().year
    return [f"FY{year}" for year in range(current_year, current_year - 16, -1)]


@app.get("/", response_class=HTMLResponse)
async def home(request: Request):
    return templates.TemplateResponse("index.html", {
        "request": request,
        "fy_choices": get_current_fy_choices(),
        "quarters": [q.value for q in Quarter],
    })


@app.get("/api/company/search")
async def api_search_company(q: str):
    if len(q) < 1:
        return []
    results = await search_companies(q)
    return results


@app.get("/api/company/{scrip_code}")
async def api_get_company(scrip_code: str):
    company = await get_company_by_scrip(scrip_code)
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    return company


@app.get("/api/validate-url")
async def api_validate_url(url: str):
    result = await validate_url(url)
    return result


@app.post("/api/submit")
async def api_submit(
    background_tasks: BackgroundTasks,
    company_name: str = Form(...),
    bse_scrip_code: str = Form(...),
    fy: str = Form(...),
    quarter: str = Form(...),
    earnings_transcript_url: str = Form(None),
    earnings_transcript_date: str = Form(None),
    earnings_transcript_html_address: str = Form(None),
    investor_presentation_url: str = Form(None),
    investor_presentation_date: str = Form(None),
    investor_presentation_html_address: str = Form(None),
    financial_results_url: str = Form(None),
    financial_results_date: str = Form(None),
    financial_results_html_address: str = Form(None),
    earnings_audio_url: str = Form(None),
    earnings_audio_date: str = Form(None),
    earnings_audio_html_address: str = Form(None),
):
    task_id = str(uuid.uuid4())[:8]

    await save_company(company_name, bse_scrip_code)

    folder_path = generate_folder_path(bse_scrip_code, fy, quarter)
    await create_download_task(task_id, company_name, bse_scrip_code, fy, quarter, folder_path)

    docs_to_process = []
    if earnings_transcript_url:
        docs_to_process.append({
            "type": "earnings_transcript",
            "url": earnings_transcript_url,
            "html_address": earnings_transcript_html_address,
            "date": earnings_transcript_date or financial_results_date or earnings_transcript_date,
            "doc_type": DocumentType.EARNINGS_CALL_TRANSCRIPT
        })
    if investor_presentation_url:
        docs_to_process.append({
            "type": "investor_presentation",
            "url": investor_presentation_url,
            "html_address": investor_presentation_html_address,
            "date": investor_presentation_date or financial_results_date or earnings_transcript_date,
            "doc_type": DocumentType.INVESTOR_PRESENTATION
        })
    if financial_results_url:
        docs_to_process.append({
            "type": "financial_results",
            "url": financial_results_url,
            "html_address": financial_results_html_address,
            "date": financial_results_date or earnings_transcript_date,
            "doc_type": DocumentType.FINANCIAL_RESULTS
        })
    # Skip audio files - they're too large to download
    # if earnings_audio_url:
    #     docs_to_process.append({
    #         "type": "earnings_audio",
    #         "url": earnings_audio_url,
    #         "date": earnings_audio_date or financial_results_date or earnings_transcript_date,
    #         "doc_type": DocumentType.EARNINGS_CALL_AUDIO
    #     })

    if docs_to_process:
        background_tasks.add_task(
            process_all_documents,
            task_id, docs_to_process, bse_scrip_code, fy, quarter, folder_path
        )
    else:
        await update_task_status(task_id, "completed")

    return {"task_id": task_id, "message": "Task submitted successfully"}


async def process_all_documents(task_id: str, docs, scrip_code: str, fy: str, quarter: str, folder_path: str):
    from database import update_file_record, update_task_status as db_update_status, get_file_records

    await db_update_status(task_id, "validating")

    for doc in docs:
        doc_date = date.today()
        if doc["date"]:
            try:
                doc_date = date.fromisoformat(doc["date"])
            except ValueError:
                pass

        filename = generate_filename(doc_date, scrip_code, doc["doc_type"], doc["url"])
        local_path = DOWNLOAD_DIR / folder_path / filename
        local_path.parent.mkdir(parents=True, exist_ok=True)

        db_record_id = await create_file_record(task_id, doc["type"], doc["url"], doc.get("html_address"), str(doc_date) if doc["date"] else None)

        await db_update_status(task_id, "downloading")

        success, error = await download_file(doc["url"], local_path)

        if success:
            await update_file_record(db_record_id, "downloaded", local_path=str(local_path))

            if S3_BUCKET:
                await db_update_status(task_id, "uploading")
                s3_key = f"{folder_path}{filename}"
                upload_success, s3_result = upload_to_s3(local_path, S3_BUCKET, s3_key)
                if upload_success:
                    await update_file_record(db_record_id, "uploaded", s3_path=s3_result)
                else:
                    await update_file_record(db_record_id, "upload_failed", error_message=s3_result)
            else:
                await update_file_record(db_record_id, "uploaded", s3_path="local_only")
        else:
            await update_file_record(db_record_id, "failed", error_message=error)

    all_records = await get_file_records(task_id)
    all_completed = all(r["status"] in ["uploaded", "upload_failed"] for r in all_records)
    any_failed = any(r["status"] in ["failed", "upload_failed"] for r in all_records)

    await db_update_status(task_id, "completed" if all_completed else "failed")


@app.get("/status/{task_id}", response_class=HTMLResponse)
async def view_status(task_id: str, request: Request):
    task = await get_task(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    files = await get_file_records(task_id)
    return templates.TemplateResponse("status.html", {
        "request": request,
        "task": task,
        "files": files
    })


@app.get("/api/status/{task_id}")
async def api_get_status(task_id: str):
    task = await get_task(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    files = await get_file_records(task_id)
    return {
        "task_id": task_id,
        "status": task["status"],
        "folder_path": task["folder_path"],
        "files": files
    }


@app.get("/api/tasks")
async def api_list_tasks():
    import aiosqlite
    DATABASE_PATH = os.path.join(os.path.dirname(__file__), "transcripts.db")
    async with aiosqlite.connect(DATABASE_PATH) as db:
        db.row_factory = aiosqlite.Row
        cursor = await db.execute(
            "SELECT task_id, company_name, bse_scrip_code, fy, quarter, status, created_at FROM download_records ORDER BY created_at DESC LIMIT 50"
        )
        rows = await cursor.fetchall()
        return [dict(row) for row in rows]


@app.get("/files", response_class=HTMLResponse)
async def list_files(request: Request):
    return templates.TemplateResponse("files.html", {"request": request})


@app.get("/api/files")
async def api_list_files(search: str = None, page: int = 1, page_size: int = 20):
    result = await get_all_files(search=search, page=page, page_size=page_size)
    return result


@app.get("/api/files/download-excel")
async def download_files_excel(search: str = None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from fastapi.responses import StreamingResponse
    import io

    result = await get_all_files(search=search, page=1, page_size=10000)

    wb = Workbook()
    ws = wb.active
    ws.title = "Downloaded Files"

    # Headers
    headers = ["Ticker Used In File name", "Company Name", "FY", "Quarter", "Document Type", "HTML Address", "Source URL", "File Path", "Event Date", "Status", "Download Date"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Data
    for row_idx, file_data in enumerate(result["files"], 2):
        ws.cell(row=row_idx, column=1, value=file_data["bse_scrip_code"]).border = thin_border
        ws.cell(row=row_idx, column=2, value=file_data["company_name"]).border = thin_border
        ws.cell(row=row_idx, column=3, value=file_data["fy"]).border = thin_border
        ws.cell(row=row_idx, column=4, value=file_data["quarter"]).border = thin_border
        ws.cell(row=row_idx, column=5, value=file_data["file_type"]).border = thin_border
        ws.cell(row=row_idx, column=6, value=file_data["html_address"] or "").border = thin_border
        ws.cell(row=row_idx, column=7, value=file_data["original_url"] or "").border = thin_border
        ws.cell(row=row_idx, column=8, value=file_data["local_path"] or "").border = thin_border
        ws.cell(row=row_idx, column=9, value=file_data["event_date"] or "").border = thin_border
        ws.cell(row=row_idx, column=10, value=file_data["status"]).border = thin_border
        ws.cell(row=row_idx, column=11, value=file_data["created_at"]).border = thin_border

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 50
    ws.column_dimensions['G'].width = 50
    ws.column_dimensions['H'].width = 50
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=downloaded_files.xlsx"}
    )


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)